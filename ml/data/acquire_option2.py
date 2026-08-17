from __future__ import annotations

import argparse
import csv
import json
import re
import ssl
import urllib.request
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import certifi

from ml.data.acquire_esci import _download
from ml.data.adapters.base import file_sha256

UCI_URL = "https://archive.ics.uci.edu/static/public/502/online+retail+ii.zip"
BTS_URL = (
    "https://transtats.bts.gov/PREZIP/"
    "Origin_and_Destination_Survey_DB1BMarket_2025_1.zip"
)
USASPENDING_URL = "https://api.usaspending.gov/api/v2/search/spending_by_award/"
TLS_CONTEXT = ssl.create_default_context(cafile=certifi.where())


def _write_jsonl(path: Path, records) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w") as output:
        for record in records:
            output.write(json.dumps(record, ensure_ascii=False) + "\n")
            count += 1
    return count


def _existing(path: Path, minimum: int) -> bool:
    if not path.exists():
        return False
    with path.open() as source:
        return sum(1 for line in source if line.strip()) >= minimum


def acquire_uci(root: Path, limit: int = 30_000) -> dict[str, Any]:
    directory = root / "uci-online-retail-ii"
    archive = directory / "online-retail-ii.zip"
    records = directory / "records.jsonl"
    directory.mkdir(parents=True, exist_ok=True)
    if not archive.exists():
        _download(UCI_URL, archive)
    if not _existing(records, limit):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("install services/api[ml] with openpyxl support") from exc
        with zipfile.ZipFile(archive) as source:
            members = [value for value in source.namelist() if value.lower().endswith(".xlsx")]
            if len(members) != 1:
                raise RuntimeError("UCI archive must contain exactly one XLSX workbook")
            workbook_path = directory / Path(members[0]).name
            if not workbook_path.exists():
                with source.open(members[0]) as input_file, workbook_path.open("wb") as output:
                    while chunk := input_file.read(1024 * 1024):
                        output.write(chunk)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        seen: set[str] = set()

        def rows():
            for sheet in workbook.worksheets:
                iterator = sheet.iter_rows(values_only=True)
                headings = [str(value).strip().lower() for value in next(iterator)]
                positions = {value: index for index, value in enumerate(headings)}
                for row in iterator:
                    invoice = str(row[positions["invoice"]] or "").strip()
                    stock_code = str(row[positions["stockcode"]] or "").strip()
                    description = str(row[positions["description"]] or "").strip()
                    quantity = row[positions["quantity"]]
                    unit_price = row[positions["price"]]
                    if (
                        not invoice
                        or invoice.lower().startswith("c")
                        or invoice in seen
                        or not stock_code
                        or not description
                    ):
                        continue
                    try:
                        if int(quantity) <= 0 or float(unit_price) <= 0:
                            continue
                    except (TypeError, ValueError):
                        continue
                    seen.add(invoice)
                    yield {
                        "invoice": invoice,
                        "stock_code": stock_code,
                        "description": description,
                        "quantity": int(quantity),
                        "unit_price": float(unit_price),
                        "country": str(row[positions["country"]] or "United Kingdom"),
                    }
                    if len(seen) >= limit:
                        return

        count = _write_jsonl(records, rows())
        workbook.close()
        if count < limit:
            raise RuntimeError(f"UCI extraction produced only {count} unique invoices")
    return {
        "version": "2009-2011",
        "source_url": UCI_URL,
        "license": "CC-BY-4.0",
        "raw_sha256": file_sha256(archive),
        "sha256": file_sha256(records),
    }


def _field(row: dict[str, Any], *names: str) -> Any:
    normalized = {key.strip().lower(): value for key, value in row.items()}
    for name in names:
        value = normalized.get(name.lower())
        if value not in (None, ""):
            return value
    raise KeyError(f"none of the source fields are present: {names}")


def acquire_db1b(root: Path, limit: int = 22_000) -> dict[str, Any]:
    directory = root / "bts-db1b"
    archive = directory / "db1b-market-2025-q1.zip"
    records = directory / "records.jsonl"
    directory.mkdir(parents=True, exist_ok=True)
    if not archive.exists():
        _download(BTS_URL, archive)
    if not _existing(records, limit):
        with zipfile.ZipFile(archive) as source:
            members = [value for value in source.namelist() if value.lower().endswith(".csv")]
            if not members:
                raise RuntimeError("DB1B archive does not contain a CSV file")
            seen: set[str] = set()

            def rows():
                with source.open(members[0]) as raw:
                    lines = (line.decode("utf-8-sig", errors="replace") for line in raw)
                    for row in csv.DictReader(lines):
                        itinerary = str(_field(row, "ItinID")).strip()
                        if itinerary in seen:
                            continue
                        try:
                            fare = float(_field(row, "MktFare", "MarketFare"))
                        except (KeyError, TypeError, ValueError):
                            continue
                        if fare <= 0:
                            continue
                        seen.add(itinerary)
                        yield {
                            "itinerary_id": itinerary,
                            "origin": str(_field(row, "Origin")).strip(),
                            "destination": str(_field(row, "Dest", "Destination")).strip(),
                            "market_fare": fare,
                            "carrier": str(
                                row.get("TkCarrier")
                                or row.get("RPCarrier")
                                or row.get("OpCarrier")
                                or "db1b_anonymized_carrier"
                            ).strip(),
                        }
                        if len(seen) >= limit:
                            return

                
            count = _write_jsonl(records, rows())
        if count < limit:
            raise RuntimeError(f"DB1B extraction produced only {count} itineraries")
    return {
        "version": "2025-Q1",
        "source_url": BTS_URL,
        "license": "US-Government-Public-Domain",
        "raw_sha256": file_sha256(archive),
        "sha256": file_sha256(records),
    }


def _post_json(url: str, payload: dict[str, Any]) -> dict[str, Any]:
    request = urllib.request.Request(
        url,
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json", "User-Agent": "ACE-public-data-pipeline/1.0"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=60, context=TLS_CONTEXT) as response:
        return json.load(response)


def acquire_usaspending(
    root: Path,
    limit: int = 15_000,
    *,
    resume_before_year: int | None = None,
) -> dict[str, Any]:
    directory = root / "usaspending-awards"
    records = directory / "records.jsonl"
    directory.mkdir(parents=True, exist_ok=True)
    if not _existing(records, limit):
        seen: set[str] = set()
        write_mode = "w"
        if resume_before_year is not None:
            if not records.exists():
                raise FileNotFoundError("cannot resume USAspending without records.jsonl")
            with records.open() as source:
                for line_number, line in enumerate(source, 1):
                    if not line.strip():
                        continue
                    try:
                        award_id = str(json.loads(line)["award_id"])
                    except (KeyError, json.JSONDecodeError) as exc:
                        raise ValueError(
                            f"invalid existing USAspending record at line {line_number}"
                        ) from exc
                    if award_id in seen:
                        raise ValueError("existing USAspending records contain duplicate awards")
                    seen.add(award_id)
            write_mode = "a"

        newest_year = resume_before_year - 1 if resume_before_year else 2025
        years = range(newest_year, 2014, -1)

        def rows():
            for year in years:
                start_date, end_date = f"{year}-01-01", f"{year}-12-31"
                page = 1
                while len(seen) < limit:
                    response = _post_json(
                        USASPENDING_URL,
                        {
                            "subawards": False,
                            "limit": 100,
                            "page": page,
                            "filters": {
                                "award_type_codes": ["A", "B", "C", "D"],
                                "time_period": [
                                    {"start_date": start_date, "end_date": end_date}
                                ],
                            },
                            "fields": [
                                "Award ID",
                                "Recipient Name",
                                "Award Amount",
                                "Description",
                            ],
                            "sort": "Award Amount",
                            "order": "desc",
                        },
                    )
                    for value in response.get("results", []):
                        award_id = str(value.get("Award ID") or "").strip()
                        recipient = str(value.get("Recipient Name") or "").strip()
                        source_description = str(value.get("Description") or "").strip()
                        try:
                            amount = float(value.get("Award Amount"))
                        except (TypeError, ValueError):
                            continue
                        if not award_id or award_id in seen or not recipient or amount <= 0:
                            continue
                        description = source_description or f"Federal contract award {award_id}"
                        seen.add(award_id)
                        yield {
                            "award_id": award_id,
                            "recipient": recipient,
                            "description": description,
                            "description_origin": (
                                "real_public"
                                if source_description
                                else "derived_from_public_award_id"
                            ),
                            "amount": amount,
                        }
                        if len(seen) >= limit:
                            return
                    if not response.get("page_metadata", {}).get("hasNext"):
                        break
                    page += 1

        added = 0
        with records.open(write_mode) as output:
            for record in rows():
                output.write(json.dumps(record, ensure_ascii=False) + "\n")
                added += 1
        count = len(seen)
        if count < limit:
            raise RuntimeError(
                f"USAspending extraction produced only {count} awards ({added} added)"
            )
    return {
        "version": "2015-2025-calendar-year-search",
        "source_url": USASPENDING_URL,
        "license": "US-Government-Public-Domain",
        "sha256": file_sha256(records),
    }


def normalize_amazon_m2(root: Path, source: Path, limit: int = 50_000) -> dict[str, Any]:
    """Normalize authenticated AIcrowd CSV files supplied by the user; only UK rows are retained."""
    products_path = source / "products_train.csv"
    sessions_path = source / "sessions_train.csv"
    if not products_path.exists() or not sessions_path.exists():
        raise FileNotFoundError(
            "Amazon-M2 requires products_train.csv and sessions_train.csv from AIcrowd"
        )
    products: dict[str, dict[str, Any]] = {}
    with products_path.open(errors="replace") as input_file:
        for row in csv.DictReader(input_file):
            if str(row.get("locale", "")).upper() != "UK":
                continue
            product_id = str(row.get("id") or "").strip()
            if product_id:
                products[product_id] = row

    def parse_items(value: str) -> list[str]:
        # Amazon-M2 stores NumPy-style arrays such as ['A' 'B'] rather than
        # Python lists with commas. Extracting quoted ASINs supports both forms.
        parsed = re.findall(r"['\"]([^'\"]+)['\"]", value)
        if not parsed and value.strip() not in {"", "[]"}:
            raise TypeError("Amazon-M2 prev_items must contain quoted product IDs")
        return parsed

    def rows():
        count = 0
        with sessions_path.open(errors="replace") as input_file:
            for index, row in enumerate(csv.DictReader(input_file)):
                if str(row.get("locale", "")).upper() != "UK":
                    continue
                previous_ids = parse_items(str(row.get("prev_items", "[]")))
                next_id = str(row.get("next_item") or "").strip()
                previous = [products.get(value) for value in previous_ids]
                next_product = products.get(next_id)
                if not previous_ids or any(value is None for value in previous) or next_product is None:
                    continue
                yield {
                    "session_id": str(row.get("session_id") or f"train-{index}"),
                    "locale": "UK",
                    "previous_product_ids": previous_ids,
                    "previous_titles": [str(value.get("title") or "") for value in previous],
                    "next_product_id": next_id,
                    "next_title": str(next_product.get("title") or ""),
                    "next_description": str(next_product.get("desc") or ""),
                    "next_price": next_product.get("price") or 50,
                    "currency": "GBP",
                }
                count += 1
                if count >= limit:
                    return

    directory = root / "amazon-m2"
    records = directory / "records.jsonl"
    count = _write_jsonl(records, rows())
    if count < limit:
        raise RuntimeError(f"Amazon-M2 extraction produced only {count} UK sessions")
    return {
        "version": "kdd-cup-2023",
        "source_url": (
            "https://www.aicrowd.com/challenges/"
            "amazon-kdd-cup-23-multilingual-recommendation-challenge"
        ),
        "license": "Apache-2.0",
        "products_sha256": file_sha256(products_path),
        "sessions_sha256": file_sha256(sessions_path),
        "sha256": file_sha256(records),
    }


def write_lock(root: Path, source_name: str, metadata: dict[str, Any]) -> dict[str, Any]:
    path = root / "source-lock.json"
    lock = json.loads(path.read_text()) if path.exists() else {"sources": {}}
    lock["acquired_at"] = datetime.now(UTC).isoformat()
    lock["sources"][source_name] = metadata
    path.write_text(json.dumps(lock, indent=2, sort_keys=True) + "\n")
    return lock


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", choices=["uci", "db1b", "usaspending", "amazon-m2"])
    parser.add_argument("--target", type=Path, default=Path("ml/data/raw/option2"))
    parser.add_argument("--amazon-source", type=Path)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--resume-before-year", type=int)
    args = parser.parse_args()
    if args.source == "uci":
        metadata = acquire_uci(args.target, args.limit or 30_000)
        name = "uci-online-retail-ii"
    elif args.source == "db1b":
        metadata = acquire_db1b(args.target, args.limit or 22_000)
        name = "bts-db1b"
    elif args.source == "usaspending":
        metadata = acquire_usaspending(
            args.target,
            args.limit or 15_000,
            resume_before_year=args.resume_before_year,
        )
        name = "usaspending-awards"
    else:
        if args.amazon_source is None:
            parser.error("--amazon-source is required for the authenticated Amazon-M2 files")
        metadata = normalize_amazon_m2(args.target, args.amazon_source, args.limit or 50_000)
        name = "amazon-m2"
    print(json.dumps(write_lock(args.target, name, metadata), indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
